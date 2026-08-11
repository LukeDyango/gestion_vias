import streamlit as st
import pandas as pd
import sqlite3
import os
import io
import json
import base64
import uuid
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageOps
import gspread
from google.oauth2.service_account import Credentials as GoogleCredentials
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage

st.set_page_config(page_title="efe · Trenes & CHILE", page_icon="🚆", layout="centered", initial_sidebar_state="collapsed")

MOBILE_CSS = """
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header[data-testid="stHeader"] {background: transparent; height: 0.5rem;}

.block-container {padding-top: 0.5rem; padding-bottom: 5.   5rem; max-width: 460px;}

/* ---- Header (blue app bar) ---- */
.app-header {
    background: #0B3B8A; color: white; display: flex; align-items: center;
    justify-content: space-between; height: 46px; padding: 0 16px; box-sizing: border-box;
}
.app-header.header-solo { border-radius: 14px; }
.app-header.header-right { border-radius: 0 14px 14px 0; }
.app-header .title { font-weight: 700; font-size: 18px; }
.app-header .right-icon { font-size: 17px; opacity: .95; }

.st-key-header_row div[data-testid="stHorizontalBlock"] { gap: 0 !important; }
.st-key-btn_back button {
    background: #0B3B8A !important; color: white !important; border: none !important;
    border-radius: 14px 0 0 14px !important; font-size: 20px !important; font-weight: 700 !important;
    height: 46px !important; padding: 0 !important; box-shadow: none !important;
}

/* ---- Generic buttons ---- */
div.stButton > button, div.stFormSubmitButton > button, div.stDownloadButton > button {
    border-radius: 10px; padding: 0.6rem 1rem; font-weight: 600; width: 100%; border: 1px solid #dcdfe4;
}

/* ---- Responsive: en pantallas angostas, columnas van una debajo de otra ---- */
@media (max-width: 480px) {
    div[data-testid="stHorizontalBlock"] { flex-wrap: wrap; }
    div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] { min-width: 100% !important; flex: 1 1 100% !important; }
    .st-key-bottom_nav div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] { min-width: 0 !important; flex: 1 1 0 !important; }
    .st-key-header_row div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] { min-width: 0 !important; }
}

/* ---- Home action cards (colored) ---- */
.st-key-btn_crear_aviso button, .st-key-btn_guardar_aviso button,
.st-key-btn_iniciar_ot button, .st-key-btn_aprobar button,
.st-key-btn_guardar_reporte button {
    background: #2FA84F !important; color: #fff !important; border: none !important;
}
.st-key-btn_generar_reporte button { background: #0B3B8A !important; color: #fff !important; border: none !important; }
.st-key-btn_add_trabajo button, .st-key-btn_add_equipo_otro button, .st-key-btn_add_material button,
.st-key-btn_add_trabajador button { background: #F1F3F6 !important; color: #0B3B8A !important; border: 1px dashed #0B3B8A !important; }
.st-key-btn_backlog_avisos button { background: #0B3B8A !important; color: #fff !important; border: none !important; text-align: left !important; }
.st-key-btn_backlog_ots button { background: #3D7DD9 !important; color: #fff !important; border: none !important; text-align: left !important; }
.st-key-btn_reportes button, .st-key-btn_backup_reportes button { background: #F1F3F6 !important; color: #333 !important; border: 1px solid #dcdfe4 !important; text-align: left !important; }
.st-key-btn_mis_avisos button { background: #0B3B8A !important; color: #fff !important; border: none !important; text-align: left !important; }
.st-key-btn_rechazar button { background: #DC3545 !important; color: #fff !important; border: none !important; }
.st-key-btn_observar button { background: #E0A458 !important; color: #fff !important; border: none !important; }
.st-key-btn_cerrar button, .st-key-btn_cerrar_ot button { background: #6c757d !important; color: #fff !important; border: none !important; }

/* ---- Pills / badges ---- */
.pill { display: inline-block; padding: 3px 12px; border-radius: 999px; font-size: 12px; font-weight: 700; }

/* ---- List rows ---- */
.aviso-row { padding: 10px 0; border-bottom: 1px solid #eee; }
.aviso-title { font-weight: 700; font-size: 15px; margin-bottom: 2px; }
.aviso-desc { color: #444; font-size: 13.5px; }
.aviso-meta { color: #888; font-size: 12px; margin-top: 4px; }

/* ---- Login ---- */
.st-key-btn_login button { background: #0B3B8A !important; color: #fff !important; border: none !important; }

/* ---- Fixed bottom nav ---- */
.st-key-bottom_nav {
    position: fixed; bottom: 0; left: 0; right: 0; margin: 0 auto; max-width: 460px;
    background: white; border-top: 1px solid #e3e5e8; padding: 4px 4px 12px 4px; z-index: 999;
}
.st-key-bottom_nav div.stButton > button {
    border: none !important; background: transparent !important; color: #6c757d;
    font-weight: 600; font-size: 12px; box-shadow: none !important; padding: 6px 2px !important;
}
.nav-brand { text-align: center; color: #b6bcc4; font-size: 11px; letter-spacing: 1px; padding-top: 2px; }
</style>
"""
st.markdown(MOBILE_CSS, unsafe_allow_html=True)

# -------------------------
# Base de datos SQLite (independiente de Prueba_app2.py)
# -------------------------
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "registros_app3.db")

def get_db():
    return sqlite3.connect(DB_PATH)

# Columnas del tercer nivel de validación, específicas por rol (EFE o Supervisor Subcontrato)
AVISO_COLUMNS = [
    "AvisoID", "Activo", "Ubicacion", "Prioridad", "Tipificacion",
    "FechaHora", "Descripcion", "Adjunto", "CreadoPor", "RolCreador",
    "Estado", "OT_Creada", "ValidacionNivel", "RolFinal", "ReporteID",
    "AprobadoSacyr", "ObsSacyr", "FechaValSacyr", "RespValSacyr",
    "AprobadoADI", "ObsADI", "FechaValADI", "RespValADI",
    "AprobadoEFE", "ObsEFE", "FechaValEFE", "RespValEFE",
    "AprobadoSubcontrato", "ObsSubcontrato", "FechaValSubcontrato", "RespValSubcontrato",
]

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS activos (
        CodigoActivo TEXT PRIMARY KEY,
        NombreActivo TEXT, Ubicacion TEXT, TipoActivo TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS avisos (
        AvisoID TEXT PRIMARY KEY,
        Activo TEXT, Ubicacion TEXT, Prioridad TEXT, Tipificacion TEXT,
        FechaHora TEXT, Descripcion TEXT, Adjunto TEXT, CreadoPor TEXT, RolCreador TEXT,
        Estado TEXT, OT_Creada INTEGER,
        ValidacionNivel TEXT, RolFinal TEXT, ReporteID TEXT,
        AprobadoSacyr INTEGER, ObsSacyr TEXT, FechaValSacyr TEXT, RespValSacyr TEXT,
        AprobadoADI INTEGER, ObsADI TEXT, FechaValADI TEXT, RespValADI TEXT,
        AprobadoEFE INTEGER, ObsEFE TEXT, FechaValEFE TEXT, RespValEFE TEXT,
        AprobadoSubcontrato INTEGER, ObsSubcontrato TEXT, FechaValSubcontrato TEXT, RespValSubcontrato TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS ots (
        OTID TEXT PRIMARY KEY,
        AvisoID TEXT, Activo TEXT, Responsable TEXT, FechaProgramada TEXT, EstadoOT TEXT
    )""")

    # Migración: agrega columnas nuevas si la tabla avisos ya existía con el esquema anterior (sin romper datos previos)
    c.execute("PRAGMA table_info(avisos)")
    cols_existentes = {row[1] for row in c.fetchall()}
    columnas_nuevas = {
        "RolFinal": "TEXT DEFAULT 'EFE'",
        "ReporteID": "TEXT DEFAULT ''",
        "AprobadoSubcontrato": "INTEGER DEFAULT 0",
        "ObsSubcontrato": "TEXT DEFAULT ''",
        "FechaValSubcontrato": "TEXT DEFAULT ''",
        "RespValSubcontrato": "TEXT DEFAULT ''",
    }
    for col, decl in columnas_nuevas.items():
        if col not in cols_existentes:
            c.execute(f"ALTER TABLE avisos ADD COLUMN {col} {decl}")

    conn.commit()
    conn.close()

def load_avisos():
    conn = get_db()
    df = pd.read_sql("SELECT * FROM avisos", conn)
    conn.close()
    for col in ["OT_Creada", "AprobadoSacyr", "AprobadoADI", "AprobadoEFE", "AprobadoSubcontrato"]:
        df[col] = df[col].astype(bool)
    df["RolFinal"] = df["RolFinal"].fillna("EFE").replace("", "EFE")
    return df

def load_activos():
    conn = get_db()
    df = pd.read_sql("SELECT * FROM activos", conn)
    conn.close()
    return df

def load_ots():
    conn = get_db()
    df = pd.read_sql("SELECT * FROM ots", conn)
    conn.close()
    return df

def save_all_avisos():
    conn = get_db()
    df = st.session_state.avisos.copy()
    for col in ["OT_Creada", "AprobadoSacyr", "AprobadoADI", "AprobadoEFE", "AprobadoSubcontrato"]:
        df[col] = df[col].astype(int)
    df.to_sql("avisos", conn, if_exists="replace", index=False)
    conn.commit()
    conn.close()

def save_all_ots():
    conn = get_db()
    st.session_state.ots.to_sql("ots", conn, if_exists="replace", index=False)
    conn.commit()
    conn.close()

# -------------------------
# Reporte Diario (Personal Terreno) - base de datos en Excel
# -------------------------
REPORTES_XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reportes_diarios.xlsx")
FOTOS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fotos_reportes")

REPORTES_SHEETS = {
    "Reportes": ["ReporteID", "Fecha", "GrupoVia", "Usuario", "Observaciones", "FechaCreacion", "Ubicacion"],
    "Trabajos": ["ReporteID", "Actividad", "KmDesde", "KmHasta", "Unidad", "Cantidad", "Hombres", "HH"],
    "Equipos": ["ReporteID", "Equipo", "Cantidad"],
    "Materiales": ["ReporteID", "Material", "Cantidad", "Estado"],
    "Asistencia": ["ReporteID", "Trabajador", "Cargo", "Estado", "HoraIngreso", "HoraSalida", "HorasExtras"],
    "Fotos": ["ReporteID", "NombreArchivo", "RutaArchivo"],
}

DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

def horas_jornada_por_fecha(fecha) -> int:
    wd = fecha.weekday()  # Lunes=0 ... Domingo=6
    if wd in (0, 1):
        return 9
    if wd in (2, 3, 4):
        return 8
    return 6

ACTIVIDADES_TRABAJO = [
    "Inspección Vía",
    "Inspección de Desviador",
    "Reemplazo de Durmientes Común",
    "Mantenimiento de geometría vía",
    "Reposición elementos enrrieladura - Eclisas",
    "Reposición elementos enrrieladura - Pernos",
    "Reposición elementos enrrieladura - Tirafondos",
    "Mantenimiento de Desviador",
    "Mantenimiento de Desvíos",
    "Mantenimiento Cruce a Nivel",
    "Mantenimiento de Cruce Peatonal",
    "Mantenimiento de Alcantarillas",
    "Mantenimiento de Cunetas",
    "Reparación/Reposición de Señalización Pare",
    "Reparación/Reposición de Señal Cruce Pito",
    "Reparación/Reposición de Baliza PK",
    "Control de vegetación y retiro de desechos",
]

UNIDADES_TRABAJO = ["ml", "un", "m3", "km"]

EQUIPOS_CATALOGO = [
    "Cortadora de Rieles", "Generador Eléctrico", "Llave de Impacto", "Motosierra", "Taladro Industrial",
    "Bandeja", "Regla Nivel Trocha", "Gatas Alza Vías", "Grupo Bateo Universal", "Perforadora de Rieles",
    "Equipo de Iluminación", "Desbrozadora", "Desbrozadora de Altura", "Arriendo Radio comunicaciones",
    "Vehículo de transporte de personal", "Camión Pluma HiRail",
]

MATERIALES_CATALOGO = [
    "Sujeciones clip", "Silla tipo Pandroll", "Silla re 115 doble pestaña riel", "Tirafondo un2",
    "Tirafondo un5", "Eclisas planchuela tipo X", "Pernos rieleros tipo X", "Perno 19x300mm",
    "Señalética Vial de Cruces a Nivel y ferroviaria", "Grasa grafitada", "Asfalto en frío en cruce a nivel",
    "Durmientes de madera impregnada", "Otro (especificar)",
]

ESTADOS_MATERIAL = ["Buen Estado", "Mal Estado"]

CARGOS_TRABAJADOR = ["Jefe de Grupo", "Ayudante", "Chofer", "P5.1", "Operario Vía", "Otro"]

GRUPOS_VIA = ["01", "02", "03", "04", "05"]

# -------------------------
# Google Sheets: histórico durable (sobrevive reinicios del hosting).
# El Excel local se mantiene igual, para las descargas del momento; Sheets es
# la fuente de verdad cuando está configurada en Secrets ([google_sheets] ...).
# -------------------------
GOOGLE_SHEETS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

@st.cache_resource(show_spinner=False)
def _cliente_sheets():
    """Conecta con el Google Sheet configurado en Secrets. Si no está configurado
    o falla la conexión, devuelve None y la app sigue funcionando solo con el Excel local."""
    try:
        creds_info = dict(st.secrets["google_sheets"]["credentials"])
        sheet_id = st.secrets["google_sheets"]["sheet_id"]
    except Exception:
        return None
    try:
        creds = GoogleCredentials.from_service_account_info(creds_info, scopes=GOOGLE_SHEETS_SCOPES)
        gc = gspread.authorize(creds)
        return gc.open_by_key(sheet_id)
    except Exception:
        return None

def _hoja_sheets(spreadsheet, nombre_hoja: str):
    headers = REPORTES_SHEETS[nombre_hoja]
    try:
        ws = spreadsheet.worksheet(nombre_hoja)
    except gspread.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title=nombre_hoja, rows=2000, cols=max(len(headers), 10))
        ws.append_row(headers)
        return ws
    if not ws.row_values(1):
        ws.append_row(headers)
    return ws

def _leer_hoja_df(nombre_hoja: str) -> pd.DataFrame:
    """Lee una hoja completa como DataFrame: desde Google Sheets si está configurado
    (fuente de verdad, sobrevive reinicios), si no desde el Excel local."""
    columnas = REPORTES_SHEETS.get(nombre_hoja, [])
    sh = _cliente_sheets()
    if sh is not None:
        try:
            ws = _hoja_sheets(sh, nombre_hoja)
            registros = ws.get_all_records()
            # Sin filas de datos, get_all_records() devuelve [] y pd.DataFrame([]) queda sin
            # columnas -- forzamos las columnas esperadas para que df["ReporteID"] no reviente.
            return pd.DataFrame(registros) if registros else pd.DataFrame(columns=columnas)
        except Exception:
            pass
    if not os.path.exists(REPORTES_XLSX_PATH):
        return pd.DataFrame(columns=columnas)
    try:
        return pd.read_excel(REPORTES_XLSX_PATH, sheet_name=nombre_hoja)
    except Exception:
        return pd.DataFrame(columns=columnas)

def _agregar_filas(nombre_hoja: str, filas: list):
    """Agrega filas a una hoja: siempre al Excel local (para las descargas del momento),
    y también a Google Sheets si está configurado (para que el histórico no se pierda
    aunque la nube reinicie el almacenamiento local)."""
    if not filas:
        return
    init_reportes_excel()
    wb = openpyxl.load_workbook(REPORTES_XLSX_PATH)
    ws_local = wb[nombre_hoja]
    for fila in filas:
        ws_local.append(fila)
    wb.save(REPORTES_XLSX_PATH)

    sh = _cliente_sheets()
    if sh is not None:
        try:
            ws = _hoja_sheets(sh, nombre_hoja)
            filas_texto = [[("" if v is None else v) for v in fila] for fila in filas]
            ws.append_rows(filas_texto, value_input_option="USER_ENTERED")
        except Exception:
            pass  # si falla Sheets, el reporte igual quedó guardado en el Excel local

def _num(val, default=0.0):
    """Convierte a número de forma segura; Sheets a veces entrega '' en celdas vacías."""
    try:
        if val is None or val == "":
            return default
        return float(val)
    except (TypeError, ValueError):
        return default

def migrate_reportes_excel():
    """Agrega hojas/columnas nuevas a un libro ya existente sin tocar los datos previos."""
    wb = openpyxl.load_workbook(REPORTES_XLSX_PATH)
    changed = False
    for sheet_name, headers in REPORTES_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            changed = True
            continue
        ws = wb[sheet_name]
        header_row = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        faltantes = [h for h in headers if h not in header_row]
        if faltantes:
            base_col = len(header_row)
            for i, h in enumerate(faltantes):
                ws.cell(row=1, column=base_col + i + 1, value=h)
            changed = True
    if changed:
        wb.save(REPORTES_XLSX_PATH)

def init_reportes_excel():
    if os.path.exists(REPORTES_XLSX_PATH):
        migrate_reportes_excel()
        return
    wb = openpyxl.Workbook()
    first_sheet = wb.active
    first = True
    for sheet_name, headers in REPORTES_SHEETS.items():
        ws = first_sheet if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        ws.append(headers)
    wb.save(REPORTES_XLSX_PATH)

def next_reporte_id():
    df = _leer_hoja_df("Reportes")
    nums = []
    if not df.empty and "ReporteID" in df.columns:
        for val in df["ReporteID"].astype(str):
            try:
                nums.append(int(val.split("-")[1]))
            except Exception:
                pass
    n = (max(nums) + 1) if nums else 1
    return f"RPT-{n:05d}"

def guardar_reporte_excel(reporte_row, trabajos_rows, equipos_rows, materiales_rows, asistencia_rows, fotos_rows):
    _agregar_filas("Reportes", [reporte_row])
    _agregar_filas("Trabajos", trabajos_rows)
    _agregar_filas("Equipos", equipos_rows)
    _agregar_filas("Materiales", materiales_rows)
    _agregar_filas("Asistencia", asistencia_rows)
    _agregar_filas("Fotos", fotos_rows)

def obtener_actividades_hoy(usuario: str) -> list:
    """Actividades registradas en el/los Reporte(s) Diario(s) de hoy para este usuario."""
    df_rep = _leer_hoja_df("Reportes")
    df_trab = _leer_hoja_df("Trabajos")
    if df_rep.empty or df_trab.empty:
        return []
    hoy = datetime.now().strftime("%Y-%m-%d")
    reportes_hoy = df_rep[(df_rep["Fecha"].astype(str) == hoy) & (df_rep["Usuario"] == usuario)]["ReporteID"].tolist()
    if not reportes_hoy:
        return []
    return df_trab[df_trab["ReporteID"].isin(reportes_hoy)]["Actividad"].dropna().unique().tolist()

def _resolver_nombre(item: dict) -> str:
    if item["nombre"] == "Otro (especificar)":
        return item.get("nombre_custom", "").strip() or "Otro"
    return item["nombre"]

# Encabezados del resumen "plano" (una fila por actividad; Equipo/Materiales/Fotos-Observación/
# Asistencia van cada uno en una sola celda, con sus varios valores separados por coma).
FLAT_HEADERS = [
    "Grupo Vía", "Jefe de Grupo", "Fecha", "Ubicación", "Trabajo Realizado", "Km Desde", "Km Hasta",
    "Unidad", "Cantidad", "Horas Trabajadas", "Equipo Utilizado", "Materiales",
    "Fotos / Observación", "Asistencia",
]
FLAT_COL_WIDTHS = [12, 18, 12, 25, 30, 10, 10, 8, 10, 14, 30, 30, 35, 35]

def _texto_equipos(equipos) -> str:
    return ", ".join(f"{e['nombre']} ({e['cantidad']:.0f})" for e in equipos) or "—"

def _texto_materiales(materiales) -> str:
    return ", ".join(
        f"{_resolver_nombre(m)} ({m['cantidad']:.0f}, {m.get('estado', '') or '—'})" for m in materiales
    ) or "—"

def _texto_fotos_obs(fotos, observaciones) -> str:
    partes = []
    if fotos:
        partes.append("Fotos: " + ", ".join(nombre for nombre, _ in fotos))
    if observaciones and str(observaciones).strip():
        partes.append(f"Obs: {observaciones}")
    return " | ".join(partes) or "—"

def _texto_asistencia(asistencia) -> str:
    return ", ".join(f"{a['nombre']} ({a['cargo']}) - {a['estado']}" for a in asistencia) or "—"

def _ajustar_anchos_columnas(ws, anchos):
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

def _tabla_pdf(data):
    if len(data) <= 1:
        data.append(["—"] * len(data[0]))
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3B8A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F3F6")]),
    ]))
    return t

def generar_excel_reporte(resumen: dict) -> bytes:
    """Genera un Excel individual (no el libro maestro): una fila por actividad, con Equipo,
    Materiales, Fotos/Observación y Asistencia condensados cada uno en una sola celda."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    equipo_txt = _texto_equipos(resumen["equipos"])
    material_txt = _texto_materiales(resumen["materiales"])
    fotos_obs_txt = _texto_fotos_obs(resumen["fotos"], resumen["observaciones"])
    asistencia_txt = _texto_asistencia(resumen["asistencia"])

    ws.append(FLAT_HEADERS)
    for t in resumen["trabajos"]:
        ws.append([
            resumen["grupo_via"], resumen["usuario"], resumen["fecha"], resumen.get("ubicacion") or "—",
            t["actividad"], t["km_desde"], t["km_hasta"], t["unidad"], t["cantidad"], t["hh"],
            equipo_txt, material_txt, fotos_obs_txt, asistencia_txt,
        ])
    _ajustar_anchos_columnas(ws, FLAT_COL_WIDTHS)

    if resumen["fotos"]:
        fila = ws.max_row + 3
        ws.cell(row=fila - 1, column=1, value="FOTOGRAFÍAS")
        for nombre, ruta in resumen["fotos"]:
            try:
                img = XLImage(ruta)
                img.width, img.height = 260, 195
                ws.add_image(img, f"A{fila}")
                fila += 13
            except Exception:
                ws.cell(row=fila, column=1, value=f"(No se pudo incrustar: {nombre})")
                fila += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def generar_pdf_reporte(resumen: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                             leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    el = []
    el.append(Paragraph(f"Reporte Diario {resumen['reporte_id']}", styles["Title"]))
    el.append(Paragraph(
        f"Grupo Vía: {resumen['grupo_via']} · Fecha: {resumen['fecha']} · Jefe de Grupo: {resumen['usuario']} · "
        f"Jornada: {resumen['horas_dia']} h/trabajador",
        styles["Normal"],
    ))
    el.append(Paragraph(f"Ubicación: {resumen.get('ubicacion') or '—'}", styles["Normal"]))
    el.append(Spacer(1, 12))

    el.append(Paragraph("Trabajos", styles["Heading2"]))
    data = [["Actividad", "Km Desde", "Km Hasta", "Unidad", "Cant.", "N° Trab.", "HH"]]
    for t in resumen["trabajos"]:
        data.append([t["actividad"], t["km_desde"], t["km_hasta"], t["unidad"], t["cantidad"], t["hombres"], t["hh"]])
    el.append(_tabla_pdf(data))
    el.append(Spacer(1, 10))

    el.append(Paragraph("Equipos", styles["Heading2"]))
    data = [["Equipo", "Cantidad"]]
    for e in resumen["equipos"]:
        data.append([e["nombre"], e["cantidad"]])
    el.append(_tabla_pdf(data))
    el.append(Spacer(1, 10))

    el.append(Paragraph("Materiales", styles["Heading2"]))
    data = [["Material", "Cantidad", "Estado"]]
    for m in resumen["materiales"]:
        data.append([_resolver_nombre(m), m["cantidad"], m.get("estado", "")])
    el.append(_tabla_pdf(data))
    el.append(Spacer(1, 10))

    el.append(Paragraph("Observaciones", styles["Heading2"]))
    el.append(Paragraph(resumen["observaciones"] or "—", styles["Normal"]))
    el.append(Spacer(1, 10))

    el.append(Paragraph("Asistencia", styles["Heading2"]))
    data = [["Trabajador", "Cargo", "Estado"]]
    for a in resumen["asistencia"]:
        data.append([a["nombre"], a["cargo"], a["estado"]])
    el.append(_tabla_pdf(data))

    if resumen["fotos"]:
        el.append(Spacer(1, 10))
        el.append(Paragraph("Fotografías", styles["Heading2"]))
        for nombre, ruta in resumen["fotos"]:
            try:
                with PILImage.open(ruta) as im:
                    w, h = im.size
                ancho = 13 * cm
                alto = ancho * h / w
                el.append(RLImage(ruta, width=ancho, height=alto))
                el.append(Spacer(1, 8))
            except Exception:
                el.append(Paragraph(f"(No se pudo incrustar: {nombre})", styles["Normal"]))

    doc.build(el)
    return buf.getvalue()

def generar_documentos_reporte(resumen: dict):
    """Genera el Excel y PDF individuales. Si una foto hace fallar la incrustación
    (formatos raros de cámara que a veces se cuelan), reintenta sin fotos para que
    el reporte —ya guardado en el libro maestro— nunca se pierda por esto."""
    try:
        excel_bytes = generar_excel_reporte(resumen)
    except Exception:
        excel_bytes = generar_excel_reporte({**resumen, "fotos": []})
    try:
        pdf_bytes = generar_pdf_reporte(resumen)
    except Exception:
        pdf_bytes = generar_pdf_reporte({**resumen, "fotos": []})
    return excel_bytes, pdf_bytes

def generar_respaldo_plano() -> bytes | None:
    """Resumen legible de TODO el histórico (todas las actividades de todos los Reportes Diarios
    guardados hasta ahora): una fila por actividad, con Equipo/Materiales/Fotos-Observación/
    Asistencia condensados cada uno en una sola celda. El libro maestro (reportes_diarios.xlsx)
    sigue normalizado en hojas separadas para Power BI; esto es solo una vista aparte para leer rápido."""
    df_rep = _leer_hoja_df("Reportes")
    df_trab = _leer_hoja_df("Trabajos")
    df_equ = _leer_hoja_df("Equipos")
    df_mat = _leer_hoja_df("Materiales")
    df_asi = _leer_hoja_df("Asistencia")
    df_fot = _leer_hoja_df("Fotos")
    if df_rep.empty:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(FLAT_HEADERS)

    for _, rep in df_rep.iterrows():
        rid = rep["ReporteID"]
        trabajos_rep = df_trab[df_trab["ReporteID"] == rid]
        if trabajos_rep.empty:
            continue

        equipo_txt = ", ".join(
            f"{r['Equipo']} ({_num(r['Cantidad']):.0f})" for _, r in df_equ[df_equ["ReporteID"] == rid].iterrows()
        ) or "—"
        material_txt = ", ".join(
            f"{r['Material']} ({_num(r['Cantidad']):.0f}, {r.get('Estado') or '—'})"
            for _, r in df_mat[df_mat["ReporteID"] == rid].iterrows()
        ) or "—"
        fotos_lista = df_fot[df_fot["ReporteID"] == rid]["NombreArchivo"].tolist()
        partes_fo = []
        if fotos_lista:
            partes_fo.append("Fotos: " + ", ".join(fotos_lista))
        obs = rep.get("Observaciones")
        if isinstance(obs, str) and obs.strip():
            partes_fo.append(f"Obs: {obs}")
        fotos_obs_txt = " | ".join(partes_fo) or "—"
        asistencia_txt = ", ".join(
            f"{r['Trabajador']} ({r['Cargo']}) - {r['Estado']}" for _, r in df_asi[df_asi["ReporteID"] == rid].iterrows()
        ) or "—"

        ubicacion = rep.get("Ubicacion")
        ubicacion_txt = ubicacion if isinstance(ubicacion, str) and ubicacion.strip() else "—"

        for _, t in trabajos_rep.iterrows():
            ws.append([
                rep["GrupoVia"], rep["Usuario"], rep["Fecha"], ubicacion_txt,
                t["Actividad"], t["KmDesde"], t["KmHasta"], t["Unidad"], t["Cantidad"], t["HH"],
                equipo_txt, material_txt, fotos_obs_txt, asistencia_txt,
            ])

    _ajustar_anchos_columnas(ws, FLAT_COL_WIDTHS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def mostrar_detalle_reporte_diario(reporte_id: str):
    """Vista de solo lectura de un Reporte Diario, para que el validador vea exactamente
    lo que ingresó Personal de Terreno (sin poder editarlo desde acá)."""
    df_rep = _leer_hoja_df("Reportes")
    df_trab = _leer_hoja_df("Trabajos")
    df_equ = _leer_hoja_df("Equipos")
    df_mat = _leer_hoja_df("Materiales")
    df_asi = _leer_hoja_df("Asistencia")
    df_fot = _leer_hoja_df("Fotos")
    if df_rep.empty:
        st.info("No se encontró el archivo de Reportes Diarios.")
        return

    fila_rep = df_rep[df_rep["ReporteID"] == reporte_id]
    if fila_rep.empty:
        st.info("Este Reporte Diario ya no está disponible (puede haberse limpiado el respaldo).")
        return
    rep = fila_rep.iloc[0]

    st.caption(f"📄 Reporte {reporte_id} · Grupo Vía {rep['GrupoVia']} · {rep['Fecha']} · Jefe de Grupo: {rep['Usuario']}")
    ubicacion = rep.get("Ubicacion")
    if isinstance(ubicacion, str) and ubicacion.strip():
        st.caption(f"📍 Ubicación: {ubicacion}")

    st.markdown("**🛠️ Trabajos**")
    trabajos_rep = df_trab[df_trab["ReporteID"] == reporte_id].drop(columns=["ReporteID"])
    if not trabajos_rep.empty:
        st.dataframe(trabajos_rep, hide_index=True, width="stretch")
    else:
        st.caption("Sin actividades registradas.")

    st.markdown("**🚜 Equipos**")
    equipos_rep = df_equ[df_equ["ReporteID"] == reporte_id].drop(columns=["ReporteID"])
    if not equipos_rep.empty:
        st.dataframe(equipos_rep, hide_index=True, width="stretch")
    else:
        st.caption("Sin equipos registrados.")

    st.markdown("**📦 Materiales**")
    materiales_rep = df_mat[df_mat["ReporteID"] == reporte_id].drop(columns=["ReporteID"])
    if not materiales_rep.empty:
        st.dataframe(materiales_rep, hide_index=True, width="stretch")
    else:
        st.caption("Sin materiales registrados.")

    st.markdown("**👷 Asistencia**")
    asistencia_rep = df_asi[df_asi["ReporteID"] == reporte_id]
    if not asistencia_rep.empty:
        st.dataframe(asistencia_rep[["Trabajador", "Cargo", "Estado"]], hide_index=True, width="stretch")
    else:
        st.caption("Sin asistencia registrada.")

    st.markdown("**📝 Observaciones**")
    obs = rep.get("Observaciones")
    st.write(obs if isinstance(obs, str) and obs.strip() else "—")

    st.markdown("**📷 Fotografías**")
    fotos_rep = df_fot[df_fot["ReporteID"] == reporte_id]
    if not fotos_rep.empty:
        cols = st.columns(2)
        for i, (_, f) in enumerate(fotos_rep.iterrows()):
            ruta = f["RutaArchivo"]
            with cols[i % 2]:
                if isinstance(ruta, str) and os.path.exists(ruta):
                    st.image(ruta, caption=f["NombreArchivo"], width="stretch")
                else:
                    st.caption(f"(No disponible: {f['NombreArchivo']})")
    else:
        st.caption("Sin fotografías.")

# -------------------------
# Perfiles / roles
# -------------------------
ROLES_VALIDADORES = ["Sacyr", "ADI (ITO)", "EFE", "Supervisor Subcontrato"]  # también son niveles de validación, en orden
PERFILES = ["Personal Terreno"] + ROLES_VALIDADORES

init_db()

def init_data():
    if "perfil" not in st.session_state:
        st.session_state.perfil = None
    if "usuario" not in st.session_state:
        st.session_state.usuario = ""
    if "login_form_counter" not in st.session_state:
        st.session_state.login_form_counter = 0
    if "page" not in st.session_state:
        st.session_state.page = "Inicio"
    if "activos" not in st.session_state:
        st.session_state.activos = load_activos()
    if "avisos" not in st.session_state:
        st.session_state.avisos = load_avisos()
    if "ots" not in st.session_state:
        st.session_state.ots = load_ots()
    if "selected_aviso" not in st.session_state:
        st.session_state.selected_aviso = None
    if "reporte_trabajos" not in st.session_state:
        st.session_state.reporte_trabajos = []
    if "reporte_equipos_catalogo" not in st.session_state:
        st.session_state.reporte_equipos_catalogo = {nombre: {"usado": False, "cantidad": 1.0} for nombre in EQUIPOS_CATALOGO}
    if "reporte_equipos_otros" not in st.session_state:
        st.session_state.reporte_equipos_otros = []
    if "equipo_otro_form_counter" not in st.session_state:
        st.session_state.equipo_otro_form_counter = 0
    if "reporte_materiales" not in st.session_state:
        st.session_state.reporte_materiales = []
    if "reporte_asistencia" not in st.session_state:
        st.session_state.reporte_asistencia = []
    if "asistencia_form_counter" not in st.session_state:
        st.session_state.asistencia_form_counter = 0
    if "reporte_fotos_counter" not in st.session_state:
        st.session_state.reporte_fotos_counter = 0
    if "rep_ubicacion_gps" not in st.session_state:
        st.session_state.rep_ubicacion_gps = ""
    if "ubicacion_texto_counter" not in st.session_state:
        st.session_state.ubicacion_texto_counter = 0
    if "aviso_form_counter" not in st.session_state:
        st.session_state.aviso_form_counter = 0
    if "last_reporte" not in st.session_state:
        st.session_state.last_reporte = None
    if "last_reporte_excel" not in st.session_state:
        st.session_state.last_reporte_excel = None
    if "last_reporte_pdf" not in st.session_state:
        st.session_state.last_reporte_pdf = None

init_data()

def next_aviso_id():
    df = st.session_state.avisos
    nums = []
    for x in df["AvisoID"].astype(str).tolist():
        try:
            nums.append(int(x.split("-")[1]))
        except Exception:
            pass
    n = (max(nums) + 1) if nums else 1
    return f"EFE-{n:05d}"

def next_ot_id():
    df = st.session_state.ots
    nums = []
    for x in df["OTID"].astype(str).tolist():
        try:
            nums.append(int(x.split("-")[1]))
        except Exception:
            pass
    n = (max(nums) + 1) if nums else 1
    return f"OT-{n:05d}"

def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M")

def new_row_id():
    return uuid.uuid4().hex[:8]

def new_trabajo_row():
    return {"id": new_row_id(), "actividad": ACTIVIDADES_TRABAJO[0], "km_desde": "", "km_hasta": "",
            "unidad": UNIDADES_TRABAJO[0], "cantidad": 0.0, "hombres": 0}

def new_material_row():
    return {"id": new_row_id(), "nombre": MATERIALES_CATALOGO[0], "nombre_custom": "", "cantidad": 0.0,
            "estado": ESTADOS_MATERIAL[0]}

def reset_reporte_form():
    st.session_state.reporte_trabajos = []
    st.session_state.reporte_equipos_catalogo = {nombre: {"usado": False, "cantidad": 1.0} for nombre in EQUIPOS_CATALOGO}
    st.session_state.reporte_equipos_otros = []
    st.session_state.reporte_materiales = []
    st.session_state.reporte_asistencia = []
    st.session_state.reporte_fotos_counter += 1
    st.session_state.rep_ubicacion_gps = ""
    st.session_state.ubicacion_texto_counter += 1

def elapsed_str(fecha_hora_str: str) -> str:
    try:
        dt = datetime.strptime(fecha_hora_str, "%Y-%m-%d %H:%M")
    except Exception:
        return ""
    delta = datetime.now() - dt
    total_min = max(int(delta.total_seconds() // 60), 0)
    h, m = divmod(total_min, 60)
    return f"{h}hr{m}" if h > 0 else f"{m}min"

# -------------------------
# UI helpers
# -------------------------
def _slug(label: str) -> str:
    return label.replace(" ", "_").replace("(", "").replace(")", "")

def badge(text, kind="info"):
    colors = {"ok": "#198754", "warn": "#E0A458", "bad": "#dc3545", "info": "#0d6efd", "muted": "#6c757d"}
    bg = colors.get(kind, "#0d6efd")
    st.markdown(f"<span class='pill' style='background:{bg};color:white;'>{text}</span>", unsafe_allow_html=True)

def priority_kind(p):
    if p in ["Alta", "Crítica"]:
        return "bad"
    if p == "Media":
        return "warn"
    return "ok"

def app_header(title: str, back_page: str | None = None, right_icon: str = ""):
    if back_page:
        with st.container(key="header_row"):
            col1, col2 = st.columns([1, 5])
            with col1:
                if st.button("←", key="btn_back"):
                    st.session_state.page = back_page
                    st.rerun()
            with col2:
                st.markdown(
                    f"""<div class="app-header header-right">
                            <div class="title">{title}</div>
                            <div class="right-icon">{right_icon}</div>
                        </div>""",
                    unsafe_allow_html=True,
                )
    else:
        st.markdown(
            f"""<div class="app-header header-solo">
                    <div class="title">{title}</div>
                    <div class="right-icon">{right_icon}</div>
                </div>""",
            unsafe_allow_html=True,
        )
    st.write("")

def perfil_bar():
    c1, c2 = st.columns([4, 1])
    with c1:
        st.caption(f"Perfil: **{st.session_state.perfil}** · {st.session_state.usuario}")
    with c2:
        if st.button("Salir", key="btn_salir"):
            st.session_state.perfil = None
            st.session_state.page = "Inicio"
            st.rerun()

def bottom_nav(items):
    active_slug = _slug(st.session_state.page)
    st.markdown(
        f"<style>.st-key-navbtn_{active_slug} button {{ color:#0B3B8A !important; }}</style>",
        unsafe_allow_html=True,
    )
    with st.container(key="bottom_nav"):
        st.markdown("<div class='nav-brand'>SACYR</div>", unsafe_allow_html=True)
        cols = st.columns(len(items))
        for col, (label, icon) in zip(cols, items):
            with col:
                if st.button(f"{icon} {label}", key=f"navbtn_{_slug(label)}"):
                    st.session_state.page = label
                    st.rerun()

def can_validate(role: str, aviso_row: dict) -> bool:
    if aviso_row["Estado"] in ["Rechazado", "Cerrado", "Validado", "OT creada"]:
        return False
    return aviso_row["ValidacionNivel"] == role

def final_cols(rol_final: str):
    """Columnas del tercer nivel de validación, según si el aviso lo cierra EFE o el Supervisor Subcontrato."""
    if rol_final == "Supervisor Subcontrato":
        return "AprobadoSubcontrato", "ObsSubcontrato", "FechaValSubcontrato", "RespValSubcontrato"
    return "AprobadoEFE", "ObsEFE", "FechaValEFE", "RespValEFE"

def move_to_next_level(aviso_id: str, current_level: str):
    df = st.session_state.avisos
    rol_final = df.loc[df["AvisoID"] == aviso_id, "RolFinal"].iloc[0] or "EFE"
    next_map = {"Sacyr": "ADI (ITO)", "ADI (ITO)": rol_final, rol_final: "Final"}
    nxt = next_map.get(current_level, "Final")
    if nxt == "Final":
        st.session_state.avisos.loc[df["AvisoID"] == aviso_id, "ValidacionNivel"] = "Final"
        st.session_state.avisos.loc[df["AvisoID"] == aviso_id, "Estado"] = "Validado"
    else:
        st.session_state.avisos.loc[df["AvisoID"] == aviso_id, "ValidacionNivel"] = nxt
        st.session_state.avisos.loc[df["AvisoID"] == aviso_id, "Estado"] = "En validación"

# -------------------------
# LOGIN
# -------------------------
def cargar_usuarios() -> dict:
    """Lee las 6 cuentas de acceso desde Secrets de Streamlit (nunca desde el código ni GitHub).
    En Streamlit Cloud: panel de la app -> Settings -> Secrets.
    En local: archivo .streamlit/secrets.toml (no se sube al repositorio)."""
    try:
        usuarios_raw = dict(st.secrets["usuarios"])
    except Exception:
        return {}
    usuarios = {}
    for datos in usuarios_raw.values():
        clave = str(datos.get("usuario", "")).strip().lower()
        if clave:
            usuarios[clave] = {
                "password": str(datos.get("password", "")),
                "nombre": datos.get("nombre", clave),
                "perfil": datos.get("perfil", "Personal Terreno"),
            }
    return usuarios

def render_login():
    app_header("Bienvenido")
    st.caption("Ingresa con tu usuario y contraseña.")
    counter = st.session_state.login_form_counter
    usuario_input = st.text_input("Usuario", key=f"login_usuario_{counter}")
    password_input = st.text_input("Contraseña", type="password", key=f"login_password_{counter}")

    if st.button("Ingresar", key="btn_login"):
        usuarios = cargar_usuarios()
        if not usuarios:
            st.error("Todavía no se configuraron los usuarios de esta app. Avisa al administrador.")
        else:
            cuenta = usuarios.get(usuario_input.strip().lower())
            if cuenta and password_input == cuenta["password"] and cuenta["perfil"] in PERFILES:
                st.session_state.perfil = cuenta["perfil"]
                st.session_state.usuario = cuenta["nombre"]
                st.session_state.page = "Inicio"
                st.rerun()
            else:
                st.session_state.login_form_counter += 1
                st.error("Usuario o contraseña incorrectos.")

    st.markdown("<div style='text-align:center;opacity:.5;padding-top:10px;'>sacyr</div>", unsafe_allow_html=True)

if st.session_state.perfil is None:
    render_login()
    st.stop()

# -------------------------
# Pantalla: Crear Aviso (común a todos los perfiles)
# -------------------------
def page_crear_aviso(title="Crear Aviso"):
    app_header(title, back_page="Inicio")
    perfil_bar()

    counter = st.session_state.aviso_form_counter
    es_terreno = st.session_state.perfil == "Personal Terreno"

    if es_terreno:
        actividades_hoy = obtener_actividades_hoy(st.session_state.usuario)
        opciones_activo = actividades_hoy + ["Otro"]
        activo_sel = st.selectbox(
            "Actividad relacionada (reporte de hoy)", opciones_activo, key=f"aviso_activo_{counter}",
            help="Si la incidencia ocurrió durante una actividad ya registrada hoy en tu Reporte Diario, selecciónala. Si no, usa 'Otro'.",
        )
        if activo_sel == "Otro":
            activo_otro = st.text_input(
                "Especifica a qué corresponde la incidencia", key=f"aviso_activo_otro_{counter}",
                placeholder="Ej: Bus MF-177 / Vía km 23...",
            )
            activo = activo_otro.strip() or "Otro"
        else:
            activo = activo_sel
    else:
        activos_df = st.session_state.activos
        activos_list = (activos_df["CodigoActivo"] + " - " + activos_df["NombreActivo"]).tolist()
        if activos_list:
            activo = st.selectbox("Activo", activos_list, key=f"aviso_activo_{counter}")
        else:
            st.caption("Aún no hay activos registrados en el catálogo.")
            activo = st.text_input("Activo", key=f"aviso_activo_{counter}", placeholder="Escribe el activo o ubicación...")

    ubicacion = st.text_input("Ubicación", placeholder="Ej: Andén sur / Túnel interior...", key=f"aviso_ubic_{counter}")
    prioridad = st.selectbox("Prioridad", ["Baja", "Media", "Alta", "Crítica"], index=2, key=f"aviso_prio_{counter}")
    tipificacion = st.selectbox("Tipificación", ["Falla eléctrica", "Falla mecánica", "Frenos", "Puertas",
                                                   "Balasto", "Agujas/Desviador", "Rieles", "Vegetación", "Otro"],
                                 key=f"aviso_tipi_{counter}")
    descripcion = st.text_area("Descripción", height=100, placeholder="Describe el hallazgo o avería...", key=f"aviso_desc_{counter}")
    adjunto = st.file_uploader("Adjuntar foto o PDF (opcional)", type=["png", "jpg", "jpeg", "pdf"], key=f"aviso_adj_{counter}")

    if st.button("Guardar", key="btn_guardar_aviso"):
        aviso_id = next_aviso_id()
        new_row = {
            "AvisoID": aviso_id, "Activo": activo, "Ubicacion": ubicacion or "—",
            "Prioridad": prioridad, "Tipificacion": tipificacion,
            "FechaHora": now_str(), "Descripcion": descripcion or "—",
            "Adjunto": adjunto.name if adjunto else "",
            "CreadoPor": st.session_state.usuario, "RolCreador": st.session_state.perfil,
            "Estado": "Nuevo", "OT_Creada": False,
            "ValidacionNivel": "Sacyr", "RolFinal": "EFE", "ReporteID": "",
            "AprobadoSacyr": False, "ObsSacyr": "", "FechaValSacyr": "", "RespValSacyr": "",
            "AprobadoADI": False, "ObsADI": "", "FechaValADI": "", "RespValADI": "",
            "AprobadoEFE": False, "ObsEFE": "", "FechaValEFE": "", "RespValEFE": "",
            "AprobadoSubcontrato": False, "ObsSubcontrato": "", "FechaValSubcontrato": "", "RespValSubcontrato": "",
        }
        st.session_state.avisos = pd.concat([st.session_state.avisos, pd.DataFrame([new_row])], ignore_index=True)
        save_all_avisos()
        st.session_state.aviso_form_counter = counter + 1
        st.success(f"Aviso {aviso_id} creado ✅")
        if st.session_state.perfil == "Personal Terreno":
            st.session_state.page = "Mis Avisos"
        else:
            st.session_state.selected_aviso = aviso_id
            st.session_state.page = "Detalle"
        st.rerun()

def render_aviso_row(row):
    st.markdown(
        f"""<div class="aviso-row">
                <div class="aviso-title">{row['Activo']}</div>
                <div class="aviso-desc">{row['Descripcion'][:90]}{'...' if len(row['Descripcion']) > 90 else ''}</div>
                <div class="aviso-meta">📄 {row['AvisoID']} · {row['FechaHora']} &nbsp; 🕐 {elapsed_str(row['FechaHora'])}</div>
            </div>""",
        unsafe_allow_html=True,
    )

# =========================================================
# Reporte Diario (Personal Terreno)
# =========================================================
def render_captura_ubicacion() -> str:
    """Ubicación donde se hace el reporte: por GPS del celular o escrita a mano. Devuelve el texto final."""
    modo = st.radio(
        "¿Cómo quieres registrar la ubicación?",
        ["📍 Usar GPS del celular", "✏️ Escribir descripción"],
        key="rep_ubicacion_modo", horizontal=True,
    )

    if modo == "📍 Usar GPS del celular":
        qp = st.query_params
        if "lat" in qp and "lon" in qp:
            st.session_state.rep_ubicacion_gps = f"{qp['lat']}, {qp['lon']} (GPS)"
            st.query_params.clear()

        if st.session_state.rep_ubicacion_gps:
            st.success(f"📍 {st.session_state.rep_ubicacion_gps}")
            if st.button("🔄 Volver a capturar", key="btn_gps_recapturar"):
                st.session_state.rep_ubicacion_gps = ""
                st.rerun()
            return st.session_state.rep_ubicacion_gps

        html_gps = """
        <div style="font-family: -apple-system, sans-serif;">
          <button id="btnGPS" style="
              width:100%; padding:0.65rem 1rem; border-radius:10px; font-weight:600;
              background:#0B3B8A; color:#fff; border:none; font-size:15px; cursor:pointer;">
            📍 Obtener mi ubicación actual
          </button>
          <p id="gpsStatus" style="font-size:12.5px; color:#888; margin-top:6px; min-height:16px;"></p>
        </div>
        <script>
        (function () {
          const btn = document.getElementById('btnGPS');
          const statusEl = document.getElementById('gpsStatus');
          btn.addEventListener('click', function () {
            if (!navigator.geolocation) {
              statusEl.textContent = 'Este navegador no soporta geolocalización. Usa la opción de texto.';
              return;
            }
            statusEl.textContent = 'Obteniendo ubicación... acepta el permiso si tu navegador lo pide.';
            navigator.geolocation.getCurrentPosition(function (pos) {
              const lat = pos.coords.latitude.toFixed(6);
              const lon = pos.coords.longitude.toFixed(6);
              const url = new URL(window.parent.location.href);
              url.searchParams.set('lat', lat);
              url.searchParams.set('lon', lon);
              window.parent.location.href = url.toString();
            }, function (err) {
              statusEl.textContent = 'No se pudo obtener la ubicación: ' + err.message + '. Usa la opción de texto.';
            }, { enableHighAccuracy: true, timeout: 10000 });
          });
        })();
        </script>
        """
        st.iframe(html_gps, height=90)
        return ""
    else:
        return st.text_input(
            "Descripción de la ubicación", key=f"rep_ubicacion_texto_{st.session_state.ubicacion_texto_counter}",
            placeholder="Ej: Km 23.400, cerca de la estación Malloco...",
        )

def render_trabajos_section():
    if not st.session_state.reporte_trabajos:
        st.caption("Aún no hay actividades agregadas.")
    for row in st.session_state.reporte_trabajos:
        rid = row["id"]
        with st.container(border=True):
            row["actividad"] = st.selectbox(
                "Actividad", ACTIVIDADES_TRABAJO,
                index=ACTIVIDADES_TRABAJO.index(row["actividad"]) if row["actividad"] in ACTIVIDADES_TRABAJO else 0,
                key=f"trab_act_{rid}",
            )
            c1, c2 = st.columns(2)
            with c1:
                row["km_desde"] = st.text_input("Kilómetro Desde", value=row["km_desde"], key=f"trab_kmd_{rid}")
            with c2:
                row["km_hasta"] = st.text_input("Kilómetro Hasta", value=row["km_hasta"], key=f"trab_kmh_{rid}")
            c3, c4 = st.columns(2)
            with c3:
                row["unidad"] = st.selectbox(
                    "Unidad", UNIDADES_TRABAJO,
                    index=UNIDADES_TRABAJO.index(row["unidad"]) if row["unidad"] in UNIDADES_TRABAJO else 0,
                    key=f"trab_uni_{rid}",
                )
            with c4:
                row["cantidad"] = st.number_input("Cantidad", min_value=0.0, value=float(row["cantidad"]), step=1.0, key=f"trab_cant_{rid}")
            row["hombres"] = st.number_input("N° Trabajadores (Hombre)", min_value=0, value=int(row["hombres"]), step=1, key=f"trab_hom_{rid}")
            if st.button("🗑 Eliminar actividad", key=f"trab_del_{rid}"):
                st.session_state.reporte_trabajos = [r for r in st.session_state.reporte_trabajos if r["id"] != rid]
                st.rerun()

def render_equipos_section():
    st.caption("Marca los equipos utilizados durante la jornada e indica la cantidad.")
    for nombre in EQUIPOS_CATALOGO:
        item = st.session_state.reporte_equipos_catalogo[nombre]
        slug = _slug(nombre)
        c1, c2, c3 = st.columns([3, 1, 1.3])
        with c1:
            st.markdown(nombre)
        with c2:
            item["usado"] = st.toggle("Usado", value=item["usado"], key=f"equipo_usado_{slug}", label_visibility="collapsed")
        with c3:
            if item["usado"]:
                item["cantidad"] = st.number_input(
                    "Cantidad", min_value=0.0, value=float(item["cantidad"] or 1.0), step=1.0,
                    key=f"equipo_cant_{slug}", label_visibility="collapsed",
                )
            else:
                st.caption("—")

    st.divider()
    st.markdown("**Otro equipo no listado**")
    counter = st.session_state.equipo_otro_form_counter
    c1, c2 = st.columns([3, 2])
    with c1:
        nombre_nuevo = st.text_input("Nombre del equipo", key=f"equipo_otro_nombre_{counter}", placeholder="Ej: Compactadora manual")
    with c2:
        cantidad_nueva = st.number_input("Cantidad", min_value=0.0, value=1.0, step=1.0, key=f"equipo_otro_cant_{counter}")
    if st.button("➕ Agregar", key="btn_add_equipo_otro"):
        nombre_limpio = nombre_nuevo.strip()
        if not nombre_limpio:
            st.warning("Escribe el nombre del equipo antes de agregar.")
        else:
            st.session_state.reporte_equipos_otros.append({"id": new_row_id(), "nombre": nombre_limpio, "cantidad": cantidad_nueva})
            st.session_state.equipo_otro_form_counter += 1
            st.rerun()

    for row in st.session_state.reporte_equipos_otros:
        rid = row["id"]
        c1, c2, c3 = st.columns([3, 2, 1])
        with c1:
            st.markdown(f"**{row['nombre']}**")
        with c2:
            st.caption(f"Cantidad: {row['cantidad']:.0f}")
        with c3:
            if st.button("🗑", key=f"equipo_otro_del_{rid}"):
                st.session_state.reporte_equipos_otros = [r for r in st.session_state.reporte_equipos_otros if r["id"] != rid]
                st.rerun()

def render_materiales_section():
    if not st.session_state.reporte_materiales:
        st.caption("Aún no hay materiales agregados.")
    for row in st.session_state.reporte_materiales:
        rid = row["id"]
        with st.container(border=True):
            c1, c2 = st.columns([3, 2])
            with c1:
                row["nombre"] = st.selectbox(
                    "Material", MATERIALES_CATALOGO,
                    index=MATERIALES_CATALOGO.index(row["nombre"]) if row["nombre"] in MATERIALES_CATALOGO else 0,
                    key=f"material_nombre_{rid}",
                )
            with c2:
                row["cantidad"] = st.number_input("Cantidad", min_value=0.0, value=float(row["cantidad"]), step=1.0, key=f"material_cant_{rid}")
            if row["nombre"] == "Otro (especificar)":
                row["nombre_custom"] = st.text_input("Especificar material", value=row["nombre_custom"], key=f"material_custom_{rid}")
            row["estado"] = st.selectbox(
                "Estado del material", ESTADOS_MATERIAL,
                index=ESTADOS_MATERIAL.index(row["estado"]) if row.get("estado") in ESTADOS_MATERIAL else 0,
                key=f"material_estado_{rid}",
            )
            if st.button("🗑 Eliminar material", key=f"material_del_{rid}"):
                st.session_state.reporte_materiales = [r for r in st.session_state.reporte_materiales if r["id"] != rid]
                st.rerun()

def render_asistencia_section():
    counter = st.session_state.asistencia_form_counter
    c1, c2 = st.columns([3, 2])
    with c1:
        nombre_nuevo = st.text_input("Nombre del trabajador", key=f"asist_nombre_{counter}", placeholder="Escribe el nombre...")
    with c2:
        cargo_nuevo = st.selectbox("Cargo", CARGOS_TRABAJADOR, key=f"asist_cargo_{counter}")

    if st.button("➕ Agregar", key="btn_add_trabajador"):
        nombre_limpio = nombre_nuevo.strip()
        if not nombre_limpio:
            st.warning("Escribe el nombre del trabajador antes de agregar.")
        else:
            st.session_state.reporte_asistencia.append({"nombre": nombre_limpio, "cargo": cargo_nuevo, "estado": "Asistió"})
            st.session_state.asistencia_form_counter += 1
            st.rerun()

    st.write("")
    if not st.session_state.reporte_asistencia:
        st.caption("Aún no hay trabajadores agregados a la asistencia.")
    for idx, a in enumerate(st.session_state.reporte_asistencia):
        c1, c2, c3 = st.columns([3, 2, 1])
        with c1:
            st.markdown(f"**{a['nombre']}**")
            st.caption(a["cargo"])
        with c2:
            a["estado"] = st.selectbox(
                "Estado", ["Asistió", "No asistió"],
                index=0 if a["estado"] == "Asistió" else 1,
                key=f"asist_estado_{idx}_{a['nombre']}", label_visibility="collapsed",
            )
        with c3:
            if st.button("🗑", key=f"asist_del_{idx}_{a['nombre']}"):
                st.session_state.reporte_asistencia.pop(idx)
                st.rerun()

def crear_aviso_desde_reporte(reporte_id: str, grupo_via: str, ubicacion: str, resumen_texto: str) -> str:
    """Al guardar un Reporte Diario, se levanta automáticamente un aviso para que lo validen
    Sacyr, la ITO y el Supervisor del Subcontrato (esta app es de uso exclusivo para Icafal)."""
    aviso_id = next_aviso_id()
    new_row = {
        "AvisoID": aviso_id, "Activo": f"Reporte Diario · Grupo Vía {grupo_via}",
        "Ubicacion": ubicacion or f"Grupo Vía {grupo_via}", "Prioridad": "Media", "Tipificacion": "Reporte Diario",
        "FechaHora": now_str(), "Descripcion": resumen_texto, "Adjunto": "",
        "CreadoPor": st.session_state.usuario, "RolCreador": "Personal Terreno",
        "Estado": "Nuevo", "OT_Creada": False,
        "ValidacionNivel": "Sacyr", "RolFinal": "Supervisor Subcontrato", "ReporteID": reporte_id,
        "AprobadoSacyr": False, "ObsSacyr": "", "FechaValSacyr": "", "RespValSacyr": "",
        "AprobadoADI": False, "ObsADI": "", "FechaValADI": "", "RespValADI": "",
        "AprobadoEFE": False, "ObsEFE": "", "FechaValEFE": "", "RespValEFE": "",
        "AprobadoSubcontrato": False, "ObsSubcontrato": "", "FechaValSubcontrato": "", "RespValSubcontrato": "",
    }
    st.session_state.avisos = pd.concat([st.session_state.avisos, pd.DataFrame([new_row])], ignore_index=True)
    save_all_avisos()
    return aviso_id

def equipos_seleccionados():
    seleccion = [
        {"nombre": nombre, "cantidad": item["cantidad"]}
        for nombre, item in st.session_state.reporte_equipos_catalogo.items()
        if item["usado"]
    ]
    seleccion += [{"nombre": o["nombre"], "cantidad": o["cantidad"]} for o in st.session_state.reporte_equipos_otros]
    return seleccion

def guardar_reporte(grupo_via, fecha_reporte, ubicacion, observaciones, fotos_subidas):
    trabajos = st.session_state.reporte_trabajos
    if not trabajos:
        return {"ok": False, "msg": "Agrega al menos una actividad en la sección Trabajos antes de guardar."}

    reporte_id = next_reporte_id()
    horas_dia = horas_jornada_por_fecha(fecha_reporte)
    fecha_str = fecha_reporte.strftime("%Y-%m-%d")
    reporte_row = [reporte_id, fecha_str, grupo_via, st.session_state.usuario, observaciones or "", now_str(), ubicacion or ""]

    trabajos_calc = [{**t, "hh": t["hombres"] * horas_dia} for t in trabajos]
    trabajos_rows = [
        [reporte_id, t["actividad"], t["km_desde"], t["km_hasta"], t["unidad"], t["cantidad"], t["hombres"], t["hh"]]
        for t in trabajos_calc
    ]
    equipos_usados = equipos_seleccionados()
    equipos_rows = [[reporte_id, e["nombre"], e["cantidad"]] for e in equipos_usados]
    materiales_rows = [
        [reporte_id, _resolver_nombre(m), m["cantidad"], m.get("estado", "")]
        for m in st.session_state.reporte_materiales
    ]
    asistencia_rows = [
        [reporte_id, a["nombre"], a["cargo"], a["estado"], "", "", ""]
        for a in st.session_state.reporte_asistencia
    ]

    fotos_guardadas = []
    if fotos_subidas:
        carpeta_fotos = os.path.join(FOTOS_DIR, reporte_id)
        os.makedirs(carpeta_fotos, exist_ok=True)
        for idx, foto in enumerate(fotos_subidas, start=1):
            nombre_final = f"foto_{idx:02d}.jpg"
            ruta = os.path.join(carpeta_fotos, nombre_final)
            try:
                # Normaliza a JPEG estándar: los celulares suelen entregar formatos
                # (MPO, HEIC, etc.) que PIL abre pero openpyxl no sabe incrustar al guardar.
                img = PILImage.open(foto)
                img = ImageOps.exif_transpose(img)
                img = img.convert("RGB")
                img.save(ruta, format="JPEG", quality=85)
            except Exception:
                nombre_final = foto.name
                ruta = os.path.join(carpeta_fotos, nombre_final)
                with open(ruta, "wb") as f:
                    f.write(foto.getbuffer())
            fotos_guardadas.append((nombre_final, ruta))
    fotos_rows = [[reporte_id, nombre, ruta] for nombre, ruta in fotos_guardadas]

    guardar_reporte_excel(reporte_row, trabajos_rows, equipos_rows, materiales_rows, asistencia_rows, fotos_rows)

    resumen = {
        "reporte_id": reporte_id, "fecha": fecha_str, "grupo_via": grupo_via,
        "usuario": st.session_state.usuario, "horas_dia": horas_dia,
        "ubicacion": ubicacion or "", "observaciones": observaciones or "",
        "trabajos": trabajos_calc,
        "equipos": equipos_usados,
        "materiales": list(st.session_state.reporte_materiales),
        "asistencia": list(st.session_state.reporte_asistencia),
        "fotos": fotos_guardadas,
        "total_hh": sum(t["hh"] for t in trabajos_calc),
    }

    resumen_texto = (
        f"Reporte Diario {reporte_id} · {fecha_str} · {len(trabajos_calc)} actividad(es), "
        f"{len(resumen['equipos'])} equipo(s), {len(resumen['materiales'])} material(es), "
        f"{len(resumen['asistencia'])} trabajador(es) en asistencia, {resumen['total_hh']:.0f} HH totales."
    )
    resumen["aviso_id"] = crear_aviso_desde_reporte(reporte_id, grupo_via, ubicacion, resumen_texto)

    excel_bytes, pdf_bytes = generar_documentos_reporte(resumen)

    return {"ok": True, "msg": f"Reporte {reporte_id} guardado ✅", "resumen": resumen,
            "excel_bytes": excel_bytes, "pdf_bytes": pdf_bytes}

def page_generar_reporte():
    app_header("Generar Reporte", back_page="Inicio")
    perfil_bar()

    st.markdown("#### Datos generales")
    c1, c2 = st.columns(2)
    with c1:
        grupo_via = st.selectbox("Grupo Vía", GRUPOS_VIA, key="rep_grupo_via")
    with c2:
        fecha_reporte = st.date_input("Fecha", datetime.now().date(), key="rep_fecha")
    horas_dia = horas_jornada_por_fecha(fecha_reporte)
    st.caption(f"📅 {DIAS_ES[fecha_reporte.weekday()]} → jornada estándar: **{horas_dia} h por trabajador** "
               f"(se usará para calcular las Horas Hombre de cada actividad).")

    st.markdown("**Ubicación**")
    ubicacion = render_captura_ubicacion()

    st.divider()
    st.markdown("#### 🛠️ Trabajos")
    render_trabajos_section()
    if st.button("➕ Agregar actividad", key="btn_add_trabajo"):
        st.session_state.reporte_trabajos.append(new_trabajo_row())
        st.rerun()

    st.divider()
    st.markdown("#### 🚜 Equipos")
    render_equipos_section()

    st.divider()
    st.markdown("#### 📦 Materiales")
    render_materiales_section()
    if st.button("➕ Agregar material", key="btn_add_material"):
        st.session_state.reporte_materiales.append(new_material_row())
        st.rerun()

    st.divider()
    st.markdown("#### 📷 Fotografías")
    fotos_subidas = st.file_uploader(
        "Adjuntar foto(s) del trabajo realizado", type=["png", "jpg", "jpeg"], accept_multiple_files=True,
        key=f"rep_fotos_{st.session_state.reporte_fotos_counter}",
    )

    st.divider()
    st.markdown("#### 📝 Observaciones")
    observaciones = st.text_area("Observaciones", key="rep_observaciones", height=100, label_visibility="collapsed",
                                  placeholder="Comentarios, novedades o información relevante de la jornada...")

    st.divider()
    st.markdown("#### 👷 Control de Asistencia")
    render_asistencia_section()

    st.divider()
    if st.button("💾 Guardar Reporte", key="btn_guardar_reporte"):
        resultado = guardar_reporte(grupo_via, fecha_reporte, ubicacion, observaciones, fotos_subidas)
        if resultado["ok"]:
            st.session_state.last_reporte = resultado["resumen"]
            st.session_state.last_reporte_excel = resultado["excel_bytes"]
            st.session_state.last_reporte_pdf = resultado["pdf_bytes"]
            reset_reporte_form()
            st.session_state.page = "Reporte Guardado"
            st.rerun()
        else:
            st.error(resultado["msg"])

def render_boton_compartir_excel(excel_bytes: bytes, reporte_id: str, resumen: dict):
    """Boton que abre el panel nativo de 'Compartir' del celular (Correo, Gmail, WhatsApp, etc.)
    con el Excel del reporte ya adjunto. No requiere configurar ningun correo ni servidor propio."""
    b64 = base64.b64encode(excel_bytes).decode("utf-8")
    file_name = f"Reporte_{reporte_id}.xlsx"
    asunto = f"Reporte Diario {reporte_id} - Grupo Via {resumen['grupo_via']} - {resumen['fecha']}"
    cuerpo = (
        f"Se adjunta el Reporte Diario {reporte_id} del Grupo Via {resumen['grupo_via']}, "
        f"correspondiente al {resumen['fecha']}, generado por {resumen['usuario']}."
    )

    html_template = """
    <div style="font-family: -apple-system, sans-serif;">
      <button id="btnCompartirReporte" style="
          width:100%; padding:0.65rem 1rem; border-radius:10px; font-weight:600;
          background:#2FA84F; color:#fff; border:none; font-size:15px; cursor:pointer;">
        📤 Enviar / Compartir Excel
      </button>
      <p id="compartirStatus" style="font-size:12.5px; color:#888; margin-top:6px; min-height:16px;"></p>
    </div>
    <script>
    (function () {
      const b64Data = "__B64DATA__";
      const fileName = __FILENAME_JSON__;
      const subject = __SUBJECT_JSON__;
      const bodyText = __BODY_JSON__;

      const btn = document.getElementById('btnCompartirReporte');
      const statusEl = document.getElementById('compartirStatus');

      btn.addEventListener('click', async function () {
        statusEl.textContent = 'Preparando archivo...';
        try {
          const byteChars = atob(b64Data);
          const byteNumbers = new Array(byteChars.length);
          for (let i = 0; i < byteChars.length; i++) {
            byteNumbers[i] = byteChars.charCodeAt(i);
          }
          const byteArray = new Uint8Array(byteNumbers);
          const file = new File([byteArray], fileName, {
            type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
          });

          if (navigator.canShare && navigator.canShare({ files: [file] })) {
            await navigator.share({ files: [file], title: subject, text: bodyText });
            statusEl.textContent = 'Compartido';
          } else {
            statusEl.textContent = 'Este navegador no permite adjuntar directo. Se abrira tu correo: adjunta el Excel a mano (boton Descargar Excel de arriba).';
            window.location.href = 'mailto:?subject=' + encodeURIComponent(subject) + '&body=' + encodeURIComponent(bodyText);
          }
        } catch (err) {
          if (err && err.name === 'AbortError') {
            statusEl.textContent = 'Cancelado.';
          } else {
            statusEl.textContent = 'No se pudo compartir: ' + (err && err.message ? err.message : err);
          }
        }
      });
    })();
    </script>
    """

    html_code = (
        html_template
        .replace("__B64DATA__", b64)
        .replace("__FILENAME_JSON__", json.dumps(file_name))
        .replace("__SUBJECT_JSON__", json.dumps(asunto))
        .replace("__BODY_JSON__", json.dumps(cuerpo))
    )
    st.iframe(html_code, height=90)

def page_reporte_guardado():
    app_header("Reporte Guardado", back_page="Inicio")
    perfil_bar()
    resumen = st.session_state.last_reporte
    if not resumen:
        st.info("No hay un reporte reciente para mostrar.")
        return

    st.success(f"Reporte **{resumen['reporte_id']}** guardado correctamente ✅")
    st.markdown("#### Resumen de lo reportado")
    c1, c2 = st.columns(2)
    with c1:
        st.metric("Actividades", len(resumen["trabajos"]))
        st.metric("Materiales", len(resumen["materiales"]))
        st.metric("Fotografías", len(resumen["fotos"]))
    with c2:
        st.metric("Equipos", len(resumen["equipos"]))
        st.metric("Asistencia", len(resumen["asistencia"]))
        st.metric("HH totales", f"{resumen['total_hh']:.0f}")
    st.caption(f"Grupo Vía {resumen['grupo_via']} · {resumen['fecha']} · Reportado por {resumen['usuario']}")

    st.divider()
    st.markdown("#### Descargar reporte")
    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            "⬇️ Descargar Excel", data=st.session_state.last_reporte_excel,
            file_name=f"Reporte_{resumen['reporte_id']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_descargar_excel", width="stretch",
        )
    with dl2:
        st.download_button(
            "⬇️ Descargar PDF", data=st.session_state.last_reporte_pdf,
            file_name=f"Reporte_{resumen['reporte_id']}.pdf", mime="application/pdf",
            key="btn_descargar_pdf", width="stretch",
        )

    st.divider()
    st.markdown("#### Enviar por correo")
    render_boton_compartir_excel(st.session_state.last_reporte_excel, resumen["reporte_id"], resumen)
    st.caption(
        "Abre el panel de Compartir del celular (Correo, Gmail, WhatsApp, etc.) con el Excel ya adjunto. "
        "Si el navegador no lo permite, se abrirá el correo sin adjunto — usa el botón Descargar Excel de arriba y adjúntalo a mano."
    )

    st.divider()
    st.info(
        f"Se generó el aviso **{resumen['aviso_id']}** en el Backlog de Avisos para su validación por "
        "Sacyr, la ITO y el Supervisor del Subcontrato."
    )
    st.caption("Esta aplicación es de uso exclusivo para Icafal.")

    st.divider()
    if st.button("🏠 Volver al Inicio", key="btn_volver_inicio_reporte"):
        st.session_state.last_reporte = None
        st.session_state.last_reporte_excel = None
        st.session_state.last_reporte_pdf = None
        st.session_state.page = "Inicio"
        st.rerun()

# =========================================================
# FLUJO: PERSONAL TERRENO (solo reporte)
# =========================================================
def terreno_inicio():
    app_header("Inicio", right_icon="💬")
    perfil_bar()
    df = st.session_state.avisos
    mios = df[df["CreadoPor"] == st.session_state.usuario]
    pend = int(mios[mios["Estado"].isin(["Nuevo", "En validación", "Observado"])].shape[0])

    st.markdown("### Bienvenido")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("➕  Crear Incidencia", key="btn_crear_aviso"):
            st.session_state.page = "Crear Incidencia"
            st.rerun()
    with c2:
        if st.button("📝  Generar Reporte", key="btn_generar_reporte"):
            st.session_state.page = "Generar Reporte"
            st.rerun()
    if st.button(f"📋  Mis Avisos · {pend} Pendientes", key="btn_mis_avisos"):
        st.session_state.page = "Mis Avisos"
        st.rerun()

def terreno_mis_avisos():
    app_header("Mis Avisos", back_page="Inicio")
    perfil_bar()
    df = st.session_state.avisos
    mios = df[df["CreadoPor"] == st.session_state.usuario].sort_values("FechaHora", ascending=False)
    if mios.empty:
        st.info("Aún no has creado avisos.")
        return
    estado_kind = {"Nuevo": "info", "En validación": "info", "Observado": "warn",
                   "Validado": "ok", "OT creada": "ok", "Cerrado": "muted", "Rechazado": "bad"}
    for _, row in mios.iterrows():
        colA, colB = st.columns([5, 1])
        with colA:
            render_aviso_row(row)
        with colB:
            badge(row["Estado"], estado_kind.get(row["Estado"], "info"))
        if row["Estado"] == "Observado":
            obs = row["ObsSacyr"] or row["ObsADI"] or row["ObsEFE"] or row.get("ObsSubcontrato", "")
            if obs:
                st.warning(f"Observación: {obs}")

def flujo_terreno():
    page = st.session_state.page
    if page == "Inicio":
        terreno_inicio()
    elif page == "Crear Incidencia":
        page_crear_aviso(title="Crear Incidencia")
    elif page == "Generar Reporte":
        page_generar_reporte()
    elif page == "Reporte Guardado":
        page_reporte_guardado()
    elif page == "Mis Avisos":
        terreno_mis_avisos()
    else:
        st.session_state.page = "Inicio"
        st.rerun()
    bottom_nav([("Inicio", "🏠"), ("Crear Incidencia", "➕"), ("Generar Reporte", "📝"), ("Mis Avisos", "📋")])

# =========================================================
# FLUJO: VALIDADORES (Sacyr / ADI (ITO) / EFE)
# =========================================================
def validador_inicio():
    app_header("Inicio", right_icon="💬")
    perfil_bar()
    dfA = st.session_state.avisos
    pend = int(dfA[~dfA["Estado"].isin(["OT creada", "Cerrado", "Rechazado"])].shape[0]) if not dfA.empty else 0
    ots_prog = int((st.session_state.ots["EstadoOT"] != "Cerrada").sum()) if not st.session_state.ots.empty else 0

    st.markdown("### Bienvenido")
    if st.button("➕  Crear Aviso", key="btn_crear_aviso"):
        st.session_state.page = "Crear Aviso"
        st.rerun()
    if st.button(f"ℹ️  Backlog de Avisos\n{pend} Pendientes", key="btn_backlog_avisos"):
        st.session_state.page = "Backlog"
        st.rerun()
    if st.button(f"🔧  Backlog de OTs\n{ots_prog} Programadas", key="btn_backlog_ots"):
        st.session_state.page = "OTs"
        st.rerun()
    if st.button("📊  Reportes (Demo)", key="btn_reportes"):
        st.info("Próximamente: dashboard de KPIs.")

    respaldo_bytes = generar_respaldo_plano()
    if respaldo_bytes:
        st.download_button(
            "💾  Descargar respaldo de Reportes Diarios",
            data=respaldo_bytes,
            file_name="reportes_diarios_respaldo.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_backup_reportes", width="stretch",
        )

    st.markdown("<div style='text-align:center;opacity:.5;padding-top:10px;'>sacyr</div>", unsafe_allow_html=True)

def validador_backlog():
    app_header("Backlog de Avisos", back_page="Inicio", right_icon="➕")
    perfil_bar()
    df = st.session_state.avisos.copy()
    tab1, tab2, tab3 = st.tabs(["Pendientes", "OT Creada", "Cerrados"])

    def render_list(filtered_df):
        if filtered_df.empty:
            st.info("No hay registros para mostrar.")
            return
        for _, row in filtered_df.iterrows():
            colA, colB = st.columns([5, 1])
            with colA:
                render_aviso_row(row)
            with colB:
                badge(row["Prioridad"], priority_kind(row["Prioridad"]))
                if st.button("Ver ›", key=f"ver_{row['AvisoID']}"):
                    st.session_state.selected_aviso = row["AvisoID"]
                    st.session_state.page = "Detalle"
                    st.rerun()

    with tab1:
        render_list(df[~df["Estado"].isin(["OT creada", "Cerrado", "Rechazado"])])
    with tab2:
        render_list(df[df["Estado"] == "OT creada"])
    with tab3:
        render_list(df[df["Estado"].isin(["Cerrado", "Rechazado"])])

def validador_detalle():
    app_header("Detalle Aviso / Crear OT", back_page="Backlog", right_icon="⋮")
    perfil_bar()
    aviso_id = st.session_state.selected_aviso
    if not aviso_id:
        st.info("Selecciona un aviso desde el Backlog primero.")
        return

    df = st.session_state.avisos
    row = df[df["AvisoID"] == aviso_id].iloc[0].to_dict()

    tiene_reporte = bool(row.get("ReporteID"))
    if tiene_reporte:
        tab_detalle, tab_reporte = st.tabs(["📋 Detalle del Aviso", "👁️ Ver Reporte Diario"])
    else:
        tab_detalle = st.container()
        tab_reporte = None

    with tab_detalle:
        estado_kind = {"Nuevo": "bad", "En validación": "info", "Observado": "warn",
                       "Validado": "ok", "OT creada": "ok", "Cerrado": "muted", "Rechazado": "bad"}
        c1, c2 = st.columns([3, 2])
        with c1:
            badge(row["Estado"], estado_kind.get(row["Estado"], "info"))
        with c2:
            st.markdown(f"<div style='text-align:right;color:#888;font-size:13px;'>ID: {row['AvisoID']}</div>", unsafe_allow_html=True)

        st.write("")
        st.markdown(
            f"""
            **Detectado:** {row['CreadoPor']} ({row['RolCreador']})
            **Activo:** {row['Activo']}
            **Ubicación:** {row['Ubicacion']}
            **Prioridad:** {row['Prioridad']}
            **Tipificación:** {row['Tipificacion']}
            """
        )
        if row["Adjunto"]:
            st.caption(f"📎 Adjunto: {row['Adjunto']}")
        if tiene_reporte:
            st.caption(f"🔗 Vinculado al Reporte Diario {row['ReporteID']} · ver pestaña 'Ver Reporte Diario'")
        st.caption(f"📄 {row['AvisoID']} · {row['FechaHora']} &nbsp; 🕐 {elapsed_str(row['FechaHora'])}")

        st.divider()

        rol_final = row.get("RolFinal") or "EFE"
        col_aprob, col_obs_name, col_fecha, col_resp = final_cols(rol_final)

        st.markdown(f"## Validación (Sacyr / ADI / {rol_final})")
        c1, c2, c3 = st.columns(3)
        with c1:
            badge("Sacyr", "ok" if row["AprobadoSacyr"] else "muted")
            st.caption(row["FechaValSacyr"] or "—")
        with c2:
            badge("ADI (ITO)", "ok" if row["AprobadoADI"] else "muted")
            st.caption(row["FechaValADI"] or "—")
        with c3:
            badge(rol_final, "ok" if row[col_aprob] else "muted")
            st.caption(row[col_fecha] or "—")

        st.write("")
        role = st.session_state.perfil
        can_do = can_validate(role, row)
        obs = st.text_area("Observación / comentario", height=80, key=f"obs_{aviso_id}_{role}")

        obs_col_map = {"Sacyr": "ObsSacyr", "ADI (ITO)": "ObsADI", rol_final: col_obs_name}

        colA, colB, colC = st.columns(3)
        with colA:
            if st.button("✅ Aprobar", disabled=not can_do, key="btn_aprobar"):
                if role == "Sacyr":
                    st.session_state.avisos.loc[df["AvisoID"] == aviso_id, ["AprobadoSacyr", "ObsSacyr", "FechaValSacyr", "RespValSacyr"]] = [True, obs, now_str(), st.session_state.usuario]
                elif role == "ADI (ITO)":
                    st.session_state.avisos.loc[df["AvisoID"] == aviso_id, ["AprobadoADI", "ObsADI", "FechaValADI", "RespValADI"]] = [True, obs, now_str(), st.session_state.usuario]
                elif role == rol_final:
                    st.session_state.avisos.loc[df["AvisoID"] == aviso_id, [col_aprob, col_obs_name, col_fecha, col_resp]] = [True, obs, now_str(), st.session_state.usuario]
                move_to_next_level(aviso_id, role)
                save_all_avisos()
                st.success("Aprobación registrada ✅")
                st.rerun()
        with colB:
            if st.button("🟨 Observar", disabled=not can_do, key="btn_observar"):
                st.session_state.avisos.loc[df["AvisoID"] == aviso_id, "Estado"] = "Observado"
                st.session_state.avisos.loc[df["AvisoID"] == aviso_id, "ValidacionNivel"] = "Sacyr"
                st.session_state.avisos.loc[df["AvisoID"] == aviso_id, obs_col_map[role]] = obs
                save_all_avisos()
                st.warning("Aviso observado. Vuelve a Sacyr para completar información.")
                st.rerun()
        with colC:
            if st.button("⛔ Rechazar", disabled=not can_do, key="btn_rechazar"):
                st.session_state.avisos.loc[df["AvisoID"] == aviso_id, "Estado"] = "Rechazado"
                st.session_state.avisos.loc[df["AvisoID"] == aviso_id, obs_col_map[role]] = obs
                save_all_avisos()
                st.error("Aviso rechazado.")
                st.rerun()

        st.divider()
        st.markdown("## Crear OT")
        df = st.session_state.avisos
        row = df[df["AvisoID"] == aviso_id].iloc[0].to_dict()

        if row["OT_Creada"]:
            st.info("Este aviso ya tiene una OT asociada. Ve a Backlog de OTs.")
        elif row["Estado"] != "Validado":
            st.info(f"⚠️ Para crear OT, el Aviso debe estar **Validado** (Sacyr → ADI → {rol_final}).")
        else:
            with st.form("form_ot"):
                responsable = st.selectbox("Responsable", ["Camila Pérez", "Jefe Mantenimiento", "Planner", "Técnico 1"])
                fecha_prog_d = st.date_input("Fecha Programada (día)", datetime.now().date())
                fecha_prog_t = st.time_input("Hora Programada", datetime.now().time().replace(second=0, microsecond=0))
                submitted_ot = st.form_submit_button("Iniciar OT", key="btn_iniciar_ot")
            if submitted_ot:
                fecha_prog = datetime.combine(fecha_prog_d, fecha_prog_t)
                ot_id = next_ot_id()
                new_ot = {"OTID": ot_id, "AvisoID": row["AvisoID"], "Activo": row["Activo"],
                          "Responsable": responsable, "FechaProgramada": fecha_prog.strftime("%Y-%m-%d %H:%M"), "EstadoOT": "Programada"}
                st.session_state.ots = pd.concat([st.session_state.ots, pd.DataFrame([new_ot])], ignore_index=True)
                st.session_state.avisos.loc[df["AvisoID"] == row["AvisoID"], "OT_Creada"] = True
                st.session_state.avisos.loc[df["AvisoID"] == row["AvisoID"], "Estado"] = "OT creada"
                save_all_avisos()
                save_all_ots()
                st.success(f"OT {ot_id} creada ✅")
                st.session_state.page = "OTs"
                st.rerun()

        st.write("")
        if st.button("✅ Cerrar Aviso", key="btn_cerrar"):
            st.session_state.avisos.loc[st.session_state.avisos["AvisoID"] == aviso_id, "Estado"] = "Cerrado"
            save_all_avisos()
            st.success("Aviso cerrado.")

    if tab_reporte is not None:
        with tab_reporte:
            st.caption("Solo lectura — esto es exactamente lo que ingresó Personal de Terreno.")
            mostrar_detalle_reporte_diario(row["ReporteID"])

def validador_ots():
    app_header("Backlog de OTs", back_page="Inicio", right_icon="📊")
    perfil_bar()
    df = st.session_state.ots.copy()
    if df.empty:
        st.info("Aún no hay OTs. Crea una desde un Aviso validado.")
        return
    estado_kind = {"Programada": "info", "En ejecución": "warn", "Cerrada": "muted"}
    for _, r in df.iterrows():
        st.markdown(f"**{r['OTID']}**")
        badge(r["EstadoOT"], estado_kind.get(r["EstadoOT"], "info"))
        st.write(f"Activo: {r['Activo']}")
        st.caption(f"Aviso: {r['AvisoID']} · Responsable: {r['Responsable']} · Programada: {r['FechaProgramada']}")
        cols = st.columns(3)
        with cols[0]:
            if st.button("▶ Ejecutar", key=f"exec_{r['OTID']}"):
                st.session_state.ots.loc[st.session_state.ots["OTID"] == r["OTID"], "EstadoOT"] = "En ejecución"
                save_all_ots()
                st.rerun()
        with cols[1]:
            if st.button("✅ Cerrar", key=f"close_{r['OTID']}"):
                st.session_state.ots.loc[st.session_state.ots["OTID"] == r["OTID"], "EstadoOT"] = "Cerrada"
                save_all_ots()
                st.rerun()
        with cols[2]:
            if st.button("🧾 Ver", key=f"seeav_{r['OTID']}"):
                st.session_state.selected_aviso = r["AvisoID"]
                st.session_state.page = "Detalle"
                st.rerun()
        st.divider()

def flujo_validador():
    page = st.session_state.page
    if page == "Inicio":
        validador_inicio()
    elif page == "Crear Aviso":
        page_crear_aviso()
    elif page == "Backlog":
        validador_backlog()
    elif page == "Detalle":
        validador_detalle()
    elif page == "OTs":
        validador_ots()
    else:
        st.session_state.page = "Inicio"
        st.rerun()
    bottom_nav([("Inicio", "🏠"), ("Crear Aviso", "➕"), ("Backlog", "📋"), ("OTs", "🔧")])

# -------------------------
# Enrutamiento por perfil
# -------------------------
if st.session_state.perfil == "Personal Terreno":
    flujo_terreno()
else:
    flujo_validador()
